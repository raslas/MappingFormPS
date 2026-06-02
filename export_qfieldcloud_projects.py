"""
export_qfieldcloud_projects.py

Fetches all QFieldCloud projects (owned by the authenticated user) together
with their collaborators and writes the result to an Excel file.

Requirements:
    pip install requests python-dotenv openpyxl

Credentials are loaded from .env in the same directory as this script.

Usage:
    python export_qfieldcloud_projects.py
"""

import os
from pathlib import Path

try:
    import requests
except ImportError:
    raise ImportError("'requests' not installed — run:  pip install requests python-dotenv openpyxl")

try:
    from dotenv import load_dotenv
except ImportError:
    raise ImportError("'python-dotenv' not installed — run:  pip install requests python-dotenv openpyxl")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise ImportError("'openpyxl' not installed — run:  pip install requests python-dotenv openpyxl")

load_dotenv(Path(__file__).parent / ".env")

BASE_URL = "https://app.qfield.cloud/api/v1"
USERNAME = os.environ.get("QFIELDCLOUD_USERNAME", "")
PASSWORD = os.environ.get("QFIELDCLOUD_PASSWORD", "")
OUTPUT   = Path(__file__).parent / "qfieldcloud_projects.xlsx"


def auth_headers(token: str) -> dict:
    return {"Authorization": f"token {token}"}


def api_get(token: str, path: str) -> dict | list:
    r = requests.get(f"{BASE_URL}{path}", headers=auth_headers(token))
    r.raise_for_status()
    return r.json()


def login() -> tuple[str, str]:
    if not USERNAME:
        raise RuntimeError("QFIELDCLOUD_USERNAME not set in .env")
    if not PASSWORD:
        raise RuntimeError("QFIELDCLOUD_PASSWORD not set in .env")

    r = requests.post(f"{BASE_URL}/auth/login/", json={"username": USERNAME, "password": PASSWORD})
    if r.status_code != 200:
        raise RuntimeError(f"Login failed [{r.status_code}]: {r.text}")
    data = r.json()
    print(f"  Logged in as: {data['username']}")
    return data["token"], data["username"]


def discover_api_endpoints(token: str) -> None:
    """Print the API root response to reveal registered endpoints."""
    r = requests.get(f"{BASE_URL}/", headers=auth_headers(token))
    print(f"\n=== API root [{r.status_code}] ===")
    if r.ok:
        try:
            import json
            print(json.dumps(r.json(), indent=2))
        except Exception:
            print(r.text[:1000])
    else:
        print(r.text[:500])
    print("=" * 40)


def get_collaborators(token: str, owner: str, project_name: str) -> list[str]:
    """Return a list of collaborator usernames for a project."""
    candidates = [
        f"/projects/{owner}/{project_name}/collaborators/",
        f"/collaborators/?project={owner}/{project_name}",
        f"/projects/{owner}/{project_name}/members/",
    ]
    for path in candidates:
        r = requests.get(f"{BASE_URL}{path}", headers=auth_headers(token))
        if r.status_code == 200:
            data = r.json()
            items = data.get("results", data) if isinstance(data, dict) else data
            names = []
            for item in items:
                if "collaborator" in item:
                    names.append(item["collaborator"])
                elif "member" in item and isinstance(item["member"], dict):
                    names.append(item["member"].get("username", ""))
                elif "username" in item:
                    names.append(item["username"])
            return [n for n in names if n]
        if r.status_code not in (404, 405):
            print(f"    WARNING: {path} → [{r.status_code}] {r.text[:200]}")
    return []


def main() -> None:
    print("Logging in...")
    token, username = login()

    discover_api_endpoints(token)

    print("\nFetching projects...")
    all_projects = api_get(token, "/projects/")
    # Filter to projects owned by the authenticated user
    projects = [p for p in all_projects if p.get("owner") == username]
    print(f"  Found {len(projects)} project(s) owned by '{username}'")

    rows: list[tuple[str, str]] = []
    for p in projects:
        name = p.get("name", "")
        print(f"  {name}...", end=" ", flush=True)
        collaborators = get_collaborators(token, username, name)
        print(f"{len(collaborators)} collaborator(s)")
        rows.append((name, ", ".join(collaborators)))

    # ── Write Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    center_align = Alignment(vertical="center", wrap_text=True)

    headers = ["Project name", "Collaborators"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    for row_idx, (name, collabs) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=name).alignment  = center_align
        ws.cell(row=row_idx, column=2, value=collabs).alignment = center_align

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")


if __name__ == "__main__":
    main()
